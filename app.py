"""
Mauritius Route Planner — Flask backend with Ollama AI Integration.

Endpoints:
    GET  /                      -> serves the HTML/JS frontend
    GET  /api/stations          -> list all stations
    POST /api/stations/manual   -> add one station {name, lat, lon}
    POST /api/stations/upload   -> add stations from an uploaded .xlsx file
    POST /api/stations/clear    -> remove all stations
    POST /api/route             -> compute displacement + real road distance/geometry for an ordered list of station names
    POST /api/download-xlsx     -> generate materials schedule with pole placement and costing
    POST /api/ai-process        -> analyze route using Ollama Qwen2.5 3B
    POST /api/ai-download-excel -> download AI-generated report as Excel
"""
import io
import os
import json
import re

import pandas as pd
import pyproj
import requests
import ollama
from flask import Flask, jsonify, render_template, request, send_file
from flask_cors import CORS
from geopy.distance import geodesic
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from shapely.geometry import LineString
from shapely.ops import transform as shapely_transform

app = Flask(__name__)
CORS(app)

@app.after_request
def after_request(response):
    response.headers.add('Access-Control-Allow-Origin', '*')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
    response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
    return response

# ---------------------------------------------------------------------------
# In-memory data store (resets when the server restarts).
# For multi-user or persistent use, swap this for a real database.
# ---------------------------------------------------------------------------
STATIONS = []

MAURITIUS_BBOX = (-20.6, 57.25, -19.9, 57.85)  # south, west, north, east
UTM_EPSG = 32740  # UTM zone 40S — accurate metre-based buffering for Mauritius

_to_utm = pyproj.Transformer.from_crs("EPSG:4326", f"EPSG:{UTM_EPSG}", always_xy=True).transform

# Equipment catalog with MUR pricing
EQUIPMENT = {
    "48F cable": {"unit": "meter", "price_mur": 850},
    "12F cable": {"unit": "meter", "price_mur": 450},
    "Suspension clamp": {"unit": "unit", "price_mur": 1200},
    "Anchor clamp - 48F cable": {"unit": "unit", "price_mur": 1800},
    "Anchor clamp - 12F cable": {"unit": "unit", "price_mur": 1500},
    "Universal Pole Bracket": {"unit": "unit", "price_mur": 800},
    "TespaBand": {"unit": "roll", "price_mur": 3500},
    "Buckles": {"unit": "set", "price_mur": 250},
}

# Labor costs (MUR per meter)
LABOR_COST_PER_M_OH = 150  # Overhead
LABOR_COST_PER_M_UG = 450  # Underground (trenching, conduit installation)

# Pole cost (MUR per pole)
POLE_COST_MUR = 8500

# Conduit cost (MUR per unit, 6m sections)
CONDUIT_COST_MUR = 1200

# Joint box cost (MUR per joint)
JOINT_BOX_COST_MUR = 3500

# Joint box installation labor cost (MUR per box)
JOINT_BOX_INSTALL_COST_MUR = 4000

# Ollama Model Configuration
OLLAMA_MODEL = "qwen2.5:3b"

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def find_station(name):
    for s in STATIONS:
        if s["name"] == name:
            return s
    return None


def fetch_road_route(coords_latlon):
    """
    Query OSRM's public routing API for a real road-following route through
    the given ordered list of (lat, lon) points.

    Returns (road_geometry, leg_distances_km, total_km).
    road_geometry is a list of [lat, lon] points tracing the actual road path,
    suitable for drawing directly on a Leaflet map.
    """
    coord_str = ";".join(f"{lon},{lat}" for lat, lon in coords_latlon)
    url = f"https://router.project-osrm.org/route/v1/driving/{coord_str}"
    response = requests.get(
        url,
        params={"overview": "full", "geometries": "geojson"},
        headers={"User-Agent": "MauritiusRoutePlanner/1.0 (educational project)"},
        timeout=20,
    )
    response.raise_for_status()
    data = response.json()

    if data.get("code") != "Ok" or not data.get("routes"):
        raise ValueError("No road route could be found between these stations.")

    route = data["routes"][0]
    road_geometry = [[lat, lon] for lon, lat in route["geometry"]["coordinates"]]
    leg_km = [leg["distance"] / 1000 for leg in route["legs"]]
    total_km = route["distance"] / 1000
    return road_geometry, leg_km, total_km


def haversine_m(lat1, lon1, lat2, lon2):
    """Calculate distance in meters between two points."""
    return geodesic((lat1, lon1), (lat2, lon2)).m


def interpolate_point(geom, cum, target_m):
    """Interpolate a point on the geometry at the given distance."""
    for i in range(len(cum) - 1):
        if cum[i + 1] >= target_m:
            frac = (target_m - cum[i]) / (cum[i + 1] - cum[i])
            lat = geom[i][0] + frac * (geom[i + 1][0] - geom[i][0])
            lon = geom[i][1] + frac * (geom[i + 1][1] - geom[i][1])
            return (lat, lon)
    return geom[-1]


def generate_materials_schedule(route_geom, ug_segments):
    """
    Generate materials schedule:
    - Every 50m: place a pole (OH section)
    - UG sections: no poles, but conduits and cables
    - Returns poles list and detailed equipment/material quantities
    """
    # Build cumulative distances
    cum = [0]
    for i in range(1, len(route_geom)):
        cum.append(cum[-1] + haversine_m(
            route_geom[i-1][0], route_geom[i-1][1],
            route_geom[i][0], route_geom[i][1]
        ))
    total_m = cum[-1]

    # Sort UG segments and merge if overlapping
    if ug_segments:
        ug_segments = sorted(ug_segments, key=lambda x: x["start_pct"])
        merged = [ug_segments[0]]
        for seg in ug_segments[1:]:
            if seg["start_pct"] <= merged[-1]["end_pct"]:
                merged[-1]["end_pct"] = max(merged[-1]["end_pct"], seg["end_pct"])
            else:
                merged.append(seg)
        ug_segments = merged

    # Determine which sections are UG
    def is_underground(pct):
        for seg in ug_segments:
            if seg["start_pct"] <= pct <= seg["end_pct"]:
                return True
        return False

    # Generate pole list (every 50m, only on OH sections)
    poles = []
    pole_spacing = 50  # meters
    current_dist = 0
    
    while current_dist < total_m:
        pct = (current_dist / total_m) * 100
        is_ug = is_underground(pct)
        
        if not is_ug:
            lat, lon = interpolate_point(route_geom, cum, current_dist)
            poles.append({
                "pole_no": len(poles) + 1,
                "distance_m": round(current_dist, 1),
                "percentage": round(pct, 2),
                "latitude": round(lat, 6),
                "longitude": round(lon, 6),
                "section_type": "Overhead"
            })
        current_dist += pole_spacing

    # Calculate OH and UG lengths
    total_oh_length = 0
    total_ug_length = 0
    
    for i in range(len(cum) - 1):
        segment_len = cum[i+1] - cum[i]
        mid_pct = ((cum[i] / total_m) * 100 + (cum[i+1] / total_m) * 100) / 2
        
        is_ug = False
        for seg in ug_segments:
            if seg["start_pct"] <= mid_pct <= seg["end_pct"]:
                is_ug = True
                break
        
        if is_ug:
            total_ug_length += segment_len
        else:
            total_oh_length += segment_len

    # Calculate equipment quantities
    pole_count = len(poles)
    
    # Cable lengths (48F for backbone, 12F for branches)
    cable_48f_m = total_m * 1.05  # 5% extra for sag/slack
    cable_12f_m = total_m * 0.15  # 15% for branches/drops
    
    # Clamps: 2 per pole for 48F, 1 per pole for 12F
    suspension_clamps = pole_count * 2
    anchor_clamps_48f = pole_count * 1
    anchor_clamps_12f = pole_count * 0.5  # Every other pole
    
    # Brackets: 1 per pole
    universal_brackets = pole_count
    
    # TespaBand: 1 roll per 10 poles
    tespa_bands = max(1, int(pole_count / 10) + 1)
    
    # Buckles: 2 per pole
    buckles = pole_count * 2
    
    # Conduit units (6m sections for UG)
    conduit_units = int(total_ug_length / 6) + 1 if total_ug_length > 0 else 0
    
    # Joint boxes: 2 per UG segment transition
    joint_boxes = len(ug_segments) * 2 if ug_segments else 0

    # Calculate costs
    equipment_costs = {
        "48F cable": cable_48f_m * EQUIPMENT["48F cable"]["price_mur"],
        "12F cable": cable_12f_m * EQUIPMENT["12F cable"]["price_mur"],
        "Suspension clamp": suspension_clamps * EQUIPMENT["Suspension clamp"]["price_mur"],
        "Anchor clamp - 48F cable": anchor_clamps_48f * EQUIPMENT["Anchor clamp - 48F cable"]["price_mur"],
        "Anchor clamp - 12F cable": anchor_clamps_12f * EQUIPMENT["Anchor clamp - 12F cable"]["price_mur"],
        "Universal Pole Bracket": universal_brackets * EQUIPMENT["Universal Pole Bracket"]["price_mur"],
        "TespaBand": tespa_bands * EQUIPMENT["TespaBand"]["price_mur"],
        "Buckles": buckles * EQUIPMENT["Buckles"]["price_mur"],
    }
    
    # Additional materials not in main equipment list
    pole_cost = pole_count * POLE_COST_MUR
    conduit_cost = conduit_units * CONDUIT_COST_MUR
    joint_box_cost = joint_boxes * JOINT_BOX_COST_MUR
    joint_box_install_cost = joint_boxes * JOINT_BOX_INSTALL_COST_MUR
    
    # Labor costs
    labor_cost = (total_oh_length * LABOR_COST_PER_M_OH) + (total_ug_length * LABOR_COST_PER_M_UG)
    
    total_equipment_cost = sum(equipment_costs.values())
    total_cost = total_equipment_cost + pole_cost + conduit_cost + joint_box_cost + joint_box_install_cost + labor_cost

    # Build equipment list for export
    equipment_list = []
    for name, qty in [
        ("48F cable", cable_48f_m),
        ("12F cable", cable_12f_m),
        ("Suspension clamp", suspension_clamps),
        ("Anchor clamp - 48F cable", anchor_clamps_48f),
        ("Anchor clamp - 12F cable", anchor_clamps_12f),
        ("Universal Pole Bracket", universal_brackets),
        ("TespaBand", tespa_bands),
        ("Buckles", buckles),
    ]:
        if qty > 0:
            equipment_list.append({
                "item": name,
                "quantity": round(qty, 1) if name.endswith("cable") else int(qty),
                "unit": EQUIPMENT[name]["unit"],
                "unit_price_mur": EQUIPMENT[name]["price_mur"],
                "total_mur": round(qty * EQUIPMENT[name]["price_mur"], 2)
            })

    # Add poles, conduit, joint boxes to equipment list
    if pole_count > 0:
        equipment_list.append({
            "item": "Wooden Pole",
            "quantity": pole_count,
            "unit": "unit",
            "unit_price_mur": POLE_COST_MUR,
            "total_mur": pole_cost
        })
    
    if conduit_units > 0:
        equipment_list.append({
            "item": "Conduit (6m sections)",
            "quantity": conduit_units,
            "unit": "unit",
            "unit_price_mur": CONDUIT_COST_MUR,
            "total_mur": conduit_cost
        })
    
    if joint_boxes > 0:
        equipment_list.append({
            "item": "Joint Box",
            "quantity": joint_boxes,
            "unit": "unit",
            "unit_price_mur": JOINT_BOX_COST_MUR,
            "total_mur": joint_box_cost
        })
        equipment_list.append({
            "item": "Joint Box Installation",
            "quantity": joint_boxes,
            "unit": "unit",
            "unit_price_mur": JOINT_BOX_INSTALL_COST_MUR,
            "total_mur": joint_box_install_cost
        })

    summary = {
        "total_distance_m": round(total_m, 1),
        "total_distance_km": round(total_m / 1000, 3),
        "oh_distance_m": round(total_oh_length, 1),
        "ug_distance_m": round(total_ug_length, 1),
        "oh_distance_km": round(total_oh_length / 1000, 3),
        "ug_distance_km": round(total_ug_length / 1000, 3),
        "pole_count": pole_count,
        "pole_spacing_m": 50,
        "cable_48f_m": round(cable_48f_m, 1),
        "cable_12f_m": round(cable_12f_m, 1),
        "conduit_units": conduit_units,
        "joint_boxes": joint_boxes,
        "labor_cost_mur": round(labor_cost, 2),
        "equipment_cost_mur": round(total_equipment_cost, 2),
        "pole_cost_mur": round(pole_cost, 2),
        "conduit_cost_mur": round(conduit_cost, 2),
        "joint_box_cost_mur": round(joint_box_cost, 2),
        "joint_box_install_cost_mur": round(joint_box_install_cost, 2),
        "total_cost_mur": round(total_cost, 2),
        "ug_segment_count": len(ug_segments),
    }

    return poles, equipment_list, summary


# ---------------------------------------------------------------------------
# Pages
# ---------------------------------------------------------------------------
@app.route("/")
def index():
    return render_template("index.html")


# ---------------------------------------------------------------------------
# Station endpoints
# ---------------------------------------------------------------------------
@app.route("/api/stations", methods=["GET"])
def get_stations():
    return jsonify(STATIONS)


@app.route("/api/stations/manual", methods=["POST"])
def add_station_manual():
    data = request.get_json(force=True)
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Station name is required."}), 400
    try:
        lat = float(data["lat"])
        lon = float(data["lon"])
    except (KeyError, TypeError, ValueError):
        return jsonify({"error": "Latitude and longitude must be numbers."}), 400

    if find_station(name):
        return jsonify({"error": f"A station named '{name}' already exists."}), 400

    STATIONS.append({"name": name, "lat": lat, "lon": lon})
    return jsonify(STATIONS)


@app.route("/api/stations/upload", methods=["POST"])
def upload_stations():
    file = request.files.get("file")
    if file is None:
        return jsonify({"error": "No file uploaded."}), 400
    try:
        df = pd.read_excel(file)
    except Exception as e:
        return jsonify({"error": f"Could not read Excel file: {e}"}), 400

    df.columns = [c.strip().lower() for c in df.columns]
    required = {"stations", "latitude", "longitude"}
    if not required.issubset(set(df.columns)):
        return jsonify({"error": f"Excel must contain columns: {sorted(required)}"}), 400

    existing_names = {s["name"] for s in STATIONS}
    added = 0
    for _, row in df.iterrows():
        name = str(row["stations"])
        if name not in existing_names:
            STATIONS.append({"name": name, "lat": float(row["latitude"]), "lon": float(row["longitude"])})
            existing_names.add(name)
            added += 1

    return jsonify({"stations": STATIONS, "added": added})


@app.route("/api/stations/clear", methods=["POST"])
def clear_stations():
    STATIONS.clear()
    return jsonify(STATIONS)


# ---------------------------------------------------------------------------
# Route + distance
# ---------------------------------------------------------------------------
@app.route("/api/route", methods=["POST"])
def compute_route():
    data = request.get_json(force=True)
    names = data.get("station_names", [])
    if len(names) < 2:
        return jsonify({"error": "Select at least two stations."}), 400

    coords = []
    for name in names:
        s = find_station(name)
        if s is None:
            return jsonify({"error": f"Station '{name}' not found."}), 400
        coords.append((s["lat"], s["lon"]))

    # Straight-line displacement, leg by leg (the "as the crow flies" distance).
    legs = []
    total_displacement_km = 0.0
    for i in range(len(names) - 1):
        d_km = geodesic(coords[i], coords[i + 1]).km
        total_displacement_km += d_km
        legs.append(
            {"from": names[i], "to": names[i + 1], "displacement_km": round(d_km, 3), "road_km": None}
        )

    # Real road-following route via OSRM.
    road_geometry = []
    total_road_km = None
    road_error = None
    try:
        road_geometry, leg_km_list, total_road_km = fetch_road_route(coords)
        for leg, leg_km in zip(legs, leg_km_list):
            leg["road_km"] = round(leg_km, 3)
    except Exception as e:
        road_error = str(e)

    return jsonify(
        {
            "legs": legs,
            "total_displacement_km": round(total_displacement_km, 3),
            "total_road_km": round(total_road_km, 3) if total_road_km is not None else None,
            "road_geometry": road_geometry,
            "road_error": road_error,
        }
    )


# ---------------------------------------------------------------------------
# Ollama AI Integration
# ---------------------------------------------------------------------------

def format_equipment_for_ai(equipment_list):
    """Format equipment list for AI prompt"""
    lines = []
    for item in equipment_list:
        lines.append(f"- {item['item']}: {item['quantity']} {item['unit']} @ MUR {item['unit_price_mur']:,.2f} = MUR {item['total_mur']:,.2f}")
    return '\n'.join(lines)


def build_ai_report_excel(station_names, summary, equipment_list, poles, ug_segments, ai_report):
    """Build Excel with AI report included"""
    from openpyxl.styles import Font, PatternFill, Border, Side
    
    HEADER_FILL = PatternFill("solid", fgColor="1A3A5C")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
    THIN = Side(style="thin", color="E2E8F0")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    
    wb = Workbook()
    
    # Sheet 1: AI Report
    ws1 = wb.active
    ws1.title = "AI Report"
    
    # Title
    ws1['A1'] = "AI-GENERATED PROJECT REPORT"
    ws1['A1'].font = Font(bold=True, size=16)
    ws1.merge_cells('A1:D1')
    
    # Route info
    ws1['A3'] = "ROUTE INFORMATION"
    ws1['A3'].font = Font(bold=True, size=12)
    ws1['B3'] = f"Stations: {', '.join(station_names)}"
    ws1['A4'] = f"Total Distance: {summary['total_distance_km']:.2f} km"
    ws1['A5'] = f"Overhead: {summary['oh_distance_km']:.2f} km"
    ws1['A6'] = f"Underground: {summary['ug_distance_km']:.2f} km"
    ws1['A7'] = f"Poles: {summary['pole_count']}"
    
    # AI Report content
    ws1['A9'] = "AI ANALYSIS REPORT"
    ws1['A9'].font = Font(bold=True, size=14)
    
    # Split report into lines and add to sheet
    lines = ai_report.split('\n')
    row = 10
    for line in lines:
        ws1.cell(row=row, column=1, value=line)
        row += 1
    
    # Adjust column width
    ws1.column_dimensions['A'].width = 100
    
    # Sheet 2: Materials
    ws2 = wb.create_sheet("Materials")
    headers = ["Item", "Quantity", "Unit", "Unit Price (MUR)", "Total (MUR)"]
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    
    for row, item in enumerate(equipment_list, 2):
        ws2.cell(row=row, column=1, value=item['item'])
        ws2.cell(row=row, column=2, value=item['quantity'])
        ws2.cell(row=row, column=3, value=item['unit'])
        ws2.cell(row=row, column=4, value=item['unit_price_mur'])
        ws2.cell(row=row, column=5, value=item['total_mur'])
    
    # Add totals
    total_row = len(equipment_list) + 2
    ws2.cell(row=total_row, column=4, value="TOTAL").font = Font(bold=True)
    ws2.cell(row=total_row, column=5, value=f"=SUM(E2:E{total_row-1})").font = Font(bold=True)
    
    # Sheet 3: Pole List
    ws3 = wb.create_sheet("Pole List")
    headers = ["Pole No", "Distance (m)", "Percentage (%)", "Latitude", "Longitude", "Section Type"]
    for col, header in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    
    for row, pole in enumerate(poles, 2):
        ws3.cell(row=row, column=1, value=pole['pole_no'])
        ws3.cell(row=row, column=2, value=pole['distance_m'])
        ws3.cell(row=row, column=3, value=pole['percentage'])
        ws3.cell(row=row, column=4, value=pole['latitude'])
        ws3.cell(row=row, column=5, value=pole['longitude'])
        ws3.cell(row=row, column=6, value=pole['section_type'])
    
    # Sheet 4: Underground Segments
    ws4 = wb.create_sheet("Underground Segments")
    if ug_segments:
        headers = ["Segment", "Start %", "End %", "Length (m)"]
        for col, header in enumerate(headers, 1):
            cell = ws4.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        
        for row, seg in enumerate(ug_segments, 2):
            start_m = seg['start_pct'] / 100 * summary['total_distance_m']
            end_m = seg['end_pct'] / 100 * summary['total_distance_m']
            ws4.cell(row=row, column=1, value=f"UG {row-1}")
            ws4.cell(row=row, column=2, value=seg['start_pct'])
            ws4.cell(row=row, column=3, value=seg['end_pct'])
            ws4.cell(row=row, column=4, value=end_m - start_m)
    
    # Set column widths
    for ws in [ws1, ws2, ws3, ws4]:
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@app.route("/api/ai-process", methods=["POST"])
def ai_process_route():
    """
    Use Ollama Qwen2.5 to process route data and generate detailed report
    """
    try:
        data = request.get_json(force=True)
        
        # Extract data from request
        route_geom = data.get("road_geometry", [])
        ug_segments = data.get("ug_segments", [])
        station_names = data.get("station_names", [])
        
        if len(route_geom) < 2:
            return jsonify({"error": "Invalid route geometry."}), 400
        
        print("\n" + "="*60)
        print("🤖 AI PROCESSING STARTED")
        print("="*60)
        print(f"📊 Route: {' → '.join(station_names)}")
        print(f"📐 Generating materials schedule...")
        
        # Generate materials schedule first
        poles, equipment_list, summary = generate_materials_schedule(route_geom, ug_segments)
        
        print(f"   - Distance: {summary['total_distance_km']:.2f} km")
        print(f"   - Poles: {summary['pole_count']}")
        print(f"   - Total Cost: MUR {summary['total_cost_mur']:,.2f}")
        print("🧠 Sending to Ollama (qwen2.5:3b)...")
        
        # Create optimized prompt - shorter but complete
        prompt = f"""Analyze this fiber optic cable route and provide a detailed report.

ROUTE:
Stations: {' → '.join(station_names)}
Total Distance: {summary['total_distance_km']:.2f} km
Overhead: {summary['oh_distance_km']:.2f} km
Underground: {summary['ug_distance_km']:.2f} km
Poles: {summary['pole_count']} (50m spacing)
UG Segments: {len(ug_segments)}

MATERIALS:
{format_equipment_for_ai(equipment_list)}

COSTS (MUR):
Total: {summary['total_cost_mur']:,.2f}
Equipment: {summary['equipment_cost_mur']:,.2f}
Labor: {summary['labor_cost_mur']:,.2f}
Poles: {summary['pole_cost_mur']:,.2f}

Provide a professional report with:
1. EXECUTIVE SUMMARY - Project overview and key metrics
2. ROUTE ANALYSIS - OH vs UG breakdown with recommendations
3. MATERIAL REQUIREMENTS - List with quantities and justification
4. COST BREAKDOWN - Detailed breakdown with insights
5. RISK ASSESSMENT - Key risks and mitigation strategies
6. RECOMMENDATIONS - Best practices and next steps
7. TIMELINE - Estimated project duration

Keep it practical and actionable. Focus on engineering insights specific to fiber optic installation in Mauritius."""

        try:
            import ollama
            
            # Call Ollama with optimized settings for speed
            response = ollama.chat(
                model=OLLAMA_MODEL,
                messages=[{
                    'role': 'user',
                    'content': prompt
                }],
                stream=False,
                options={
                    'temperature': 0.3,     # Lower = more focused, faster
                    'top_k': 40,
                    'top_p': 0.9,
                    'num_predict': 3000,    # Allow enough tokens for complete report
                }
            )
            
            ai_report = response['message']['content']
            print(f"✅ AI Response received! ({len(ai_report)} characters)")
            
        except ImportError:
            print("❌ Ollama not installed!")
            ai_report = generate_fallback_report(summary, equipment_list, station_names)
        except Exception as e:
            print(f"❌ Ollama Error: {str(e)}")
            ai_report = generate_fallback_report(summary, equipment_list, station_names)
        
        print("="*60)
        print("✅ AI PROCESSING COMPLETE")
        print("="*60 + "\n")
        
        return jsonify({
            "success": True,
            "report": ai_report,
            "summary": summary,
            "equipment": equipment_list,
            "poles": poles,
            "excel_ready": True
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"AI processing failed: {str(e)}"}), 500

def generate_fallback_report(summary, equipment_list, station_names):
    """Generate a quick fallback report when Ollama is unavailable"""
    report = f"""FIBER OPTIC ROUTE ANALYSIS REPORT

1. EXECUTIVE SUMMARY
Route: {' → '.join(station_names)}
Total Distance: {summary['total_distance_km']:.2f} km
Poles Required: {summary['pole_count']}
Total Estimated Cost: MUR {summary['total_cost_mur']:,.2f}

2. ROUTE ANALYSIS
- Overhead Distance: {summary['oh_distance_km']:.2f} km ({summary['oh_distance_km']/summary['total_distance_km']*100:.1f}%)
- Underground Distance: {summary['ug_distance_km']:.2f} km ({summary['ug_distance_km']/summary['total_distance_km']*100:.1f}%)
- Pole Spacing: 50m
- UG Segments: {len(ug_segments) if 'ug_segments' in locals() else 0}

3. MATERIAL REQUIREMENTS
{format_equipment_for_ai(equipment_list)}

4. COST BREAKDOWN
- Equipment: MUR {summary['equipment_cost_mur']:,.2f} ({summary['equipment_cost_mur']/summary['total_cost_mur']*100:.1f}%)
- Labor: MUR {summary['labor_cost_mur']:,.2f} ({summary['labor_cost_mur']/summary['total_cost_mur']*100:.1f}%)
- Poles: MUR {summary['pole_cost_mur']:,.2f} ({summary['pole_cost_mur']/summary['total_cost_mur']*100:.1f}%)
- Conduit: MUR {summary['conduit_cost_mur']:,.2f}
- Joint Boxes: MUR {summary['joint_box_cost_mur']:,.2f}
- TOTAL: MUR {summary['total_cost_mur']:,.2f}

5. RECOMMENDATIONS
- Install poles at 50m intervals on overhead sections
- Use HDPE conduits for underground sections
- Include 15% material contingency
- Plan for proper grounding at each pole
- Consider cyclone-resistant hardware for overhead sections

6. PROJECT TIMELINE
Estimated duration: {max(1, int(summary['total_distance_km'] * 1.5))} working days
- Survey & Planning: {max(1, int(summary['total_distance_km'] * 0.2))} days
- Material Procurement: 3-5 days
- Construction: {max(1, int(summary['total_distance_km'] * 1.0))} days
- Testing & Commissioning: {max(1, int(summary['total_distance_km'] * 0.1))} days

7. RISK ASSESSMENT
- Weather: Cyclones can cause delays during summer months
- Right-of-Way: Permits needed for urban areas
- Soil Conditions: Rock may require specialized trenching equipment
- Traffic Management: Required for road crossings
- Material Security: Secure storage needed to prevent theft

Note: This is a calculated report. Install Ollama for AI-powered insights."""
    
    return report


@app.route("/api/ai-download-excel", methods=["POST"])
def ai_download_excel():
    """Download the AI-generated report as Excel"""
    data = request.get_json(force=True)
    
    station_names = data.get("station_names", [])
    summary = data.get("summary", {})
    equipment_list = data.get("equipment", [])
    poles = data.get("poles", [])
    ug_segments = data.get("ug_segments", [])
    ai_report = data.get("report", "")
    
    output = build_ai_report_excel(
        station_names=station_names,
        summary=summary,
        equipment_list=equipment_list,
        poles=poles,
        ug_segments=ug_segments,
        ai_report=ai_report
    )
    
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"AI_Project_Report_{station_names[0]}_{station_names[-1]}.xlsx"
    )


# ---------------------------------------------------------------------------
# Costing workbook builder (InventoryA catalog + formula-driven Cost Estimate)
# ---------------------------------------------------------------------------
NAVY = "1A3A5C"
GREY = "E2E8F0"
ICE = "EEF4FB"
INPUT_FILL = PatternFill("solid", fgColor="FFF8E1")
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
TOTAL_FILL = PatternFill("solid", fgColor=ICE)
GRAND_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=13, color=NAVY)
SUB_FONT = Font(italic=True, size=9, color="475569")
THIN = Side(style="thin", color=GREY)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Frequency rules: (item, category, unit, unit_price, frequency_basis_label, frequency_value, notes)
INVENTORY_A = [
    ("Wooden Pole",                       "Material", "unit", POLE_COST_MUR,                      "Per OH metres",   50,   "1 pole every 50m of overhead distance"),
    ("48F Cable",                         "Material", "m",    EQUIPMENT["48F cable"]["price_mur"], "Per total metre", 1.05, "Total route distance + 5% slack/sag"),
    ("12F Cable",                         "Material", "m",    EQUIPMENT["12F cable"]["price_mur"], "Per total metre", 0.15, "Total route distance x 15% for branches/drops"),
    ("Suspension Clamp",                  "Material", "unit", EQUIPMENT["Suspension clamp"]["price_mur"], "Per pole", 2,    "2 clamps per pole"),
    ("Anchor Clamp - 48F Cable",          "Material", "unit", EQUIPMENT["Anchor clamp - 48F cable"]["price_mur"], "Per pole", 1, "1 per pole"),
    ("Anchor Clamp - 12F Cable",          "Material", "unit", EQUIPMENT["Anchor clamp - 12F cable"]["price_mur"], "Per pole", 0.5, "1 per 2 poles"),
    ("Universal Pole Bracket",            "Material", "unit", EQUIPMENT["Universal Pole Bracket"]["price_mur"], "Per pole", 1, "1 per pole"),
    ("TespaBand",                         "Material", "roll", EQUIPMENT["TespaBand"]["price_mur"], "Per 10 poles", 1,   "1 roll per 10 poles (+1 roll minimum)"),
    ("Buckles",                           "Material", "set",  EQUIPMENT["Buckles"]["price_mur"],   "Per pole",     2,   "2 sets per pole"),
    ("Conduit (6m section)",              "Material", "unit", CONDUIT_COST_MUR,                    "Per UG metres", 6,  "1 section every 6m of underground distance (+1)"),
    ("Joint Box",                         "Material", "unit", JOINT_BOX_COST_MUR,                  "Per UG segment", 2, "2 joint boxes per underground segment transition"),
    ("Joint Box Installation",            "Labour",   "unit", 4000,                                 "Per UG segment", 2, "Installation labour, 2 per UG segment"),
    ("Overhead Installation Labour",      "Labour",   "m",    LABOR_COST_PER_M_OH,                  "Per OH metre",  1,  "Labour rate per metre of overhead cable run"),
    ("Underground Installation Labour",   "Labour",   "m",    LABOR_COST_PER_M_UG,                  "Per UG metre",  1,  "Labour rate per metre of underground/trenching run"),
]
INV_FIRST_ROW = 5
INV_LAST_ROW = INV_FIRST_ROW + len(INVENTORY_A) - 1


def _build_inventory_sheet(wb):
    inv = wb.active
    inv.title = "InventoryA"

    inv["A1"] = "INVENTORY A — Materials Catalog & Frequency Rules"
    inv["A1"].font = TITLE_FONT
    inv.merge_cells("A1:G1")
    inv["A2"] = "Master reference: unit pricing + usage frequency per metre/pole/segment. Cost Estimate sheet pulls all prices from here."
    inv["A2"].font = SUB_FONT
    inv.merge_cells("A2:G2")

    headers = ["Item", "Category", "Unit", "Unit Price (Rs)", "Frequency Basis", "Frequency Value", "Notes"]
    for i, h in enumerate(headers, start=1):
        c = inv.cell(row=4, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER

    r = INV_FIRST_ROW
    for item, cat, unit, price, basis, freq, notes in INVENTORY_A:
        inv.cell(row=r, column=1, value=item).border = BORDER
        cat_cell = inv.cell(row=r, column=2, value=cat)
        cat_cell.border = BORDER
        cat_cell.font = Font(color="2E6DA4" if cat == "Material" else "E65100", size=9, bold=True)
        inv.cell(row=r, column=3, value=unit).border = BORDER
        pc = inv.cell(row=r, column=4, value=price); pc.border = BORDER; pc.number_format = '#,##0'
        inv.cell(row=r, column=5, value=basis).border = BORDER
        inv.cell(row=r, column=6, value=freq).border = BORDER
        nc = inv.cell(row=r, column=7, value=notes); nc.border = BORDER; nc.font = Font(size=9, color="475569")
        r += 1

    for i, w in enumerate([28, 11, 8, 14, 16, 14, 46], start=1):
        inv.column_dimensions[get_column_letter(i)].width = w
    inv.freeze_panes = "A5"
    return inv


def _build_cost_estimate_sheet(wb, oh_m, ug_m, seg_count):
    ce = wb.create_sheet("Cost Estimate")

    ce["A1"] = "ROUTE COST ESTIMATE"
    ce["A1"].font = TITLE_FONT
    ce.merge_cells("A1:G1")
    ce["A2"] = "Distances pulled directly from the plotted route. All quantities & prices reference InventoryA live."
    ce["A2"].font = SUB_FONT
    ce.merge_cells("A2:G2")

    ce["A4"] = "ROUTE INPUTS"
    ce["A4"].font = Font(bold=True, size=11, color=NAVY)
    ce.merge_cells("A4:B4")

    input_rows = [
        ("Overhead (OH) Distance (m)", round(oh_m, 1)),
        ("Underground (UG) Distance (m)", round(ug_m, 1)),
        ("Number of UG Segments", seg_count),
    ]
    r = 5
    for label, val in input_rows:
        ce.cell(row=r, column=1, value=label).font = Font(size=10)
        c = ce.cell(row=r, column=2, value=val)
        c.fill = INPUT_FILL
        c.border = BORDER
        c.number_format = '#,##0.0'
        r += 1

    OH_CELL, UG_CELL, SEG_CELL = "$B$5", "$B$6", "$B$7"

    ce["A9"] = "Total Route Distance (m)"
    ce["A9"].font = Font(size=10)
    ce["B9"] = f"={OH_CELL}+{UG_CELL}"
    ce["B9"].number_format = '#,##0.0'
    ce["B9"].border = BORDER

    ce["A10"] = "Pole Count (1 per 50m OH)"
    ce["A10"].font = Font(size=10)
    ce["B10"] = (f"=ROUNDDOWN({OH_CELL}/INDEX(InventoryA!$F${INV_FIRST_ROW}:$F${INV_LAST_ROW},"
                 f"MATCH(\"Wooden Pole\",InventoryA!$A${INV_FIRST_ROW}:$A${INV_LAST_ROW},0)),0)")
    ce["B10"].border = BORDER
    POLE_CELL, TOTAL_CELL = "$B$10", "$B$9"

    r = 12
    ce.cell(row=r, column=1, value="MATERIALS & LABOUR (live-referenced from InventoryA)").font = Font(bold=True, size=11, color=NAVY)
    ce.merge_cells(f"A{r}:G{r}")
    table_start = r + 1
    headers = ["Item", "Category", "Unit", "Unit Price (Rs)", "Qty Formula (basis)", "Qty", "Subtotal (Rs)"]
    for i, h in enumerate(headers, start=1):
        c = ce.cell(row=table_start, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER

    def lookup(col):
        return (f'INDEX(InventoryA!${col}${INV_FIRST_ROW}:${col}${INV_LAST_ROW},'
                f'MATCH("{{item}}",InventoryA!$A${INV_FIRST_ROW}:$A${INV_LAST_ROW},0))')

    items_qty = {
        "Wooden Pole": f"={POLE_CELL}",
        "48F Cable": f"={TOTAL_CELL}*{lookup('F').format(item='48F Cable')}",
        "12F Cable": f"={TOTAL_CELL}*{lookup('F').format(item='12F Cable')}",
        "Suspension Clamp": f"={POLE_CELL}*{lookup('F').format(item='Suspension Clamp')}",
        "Anchor Clamp - 48F Cable": f"={POLE_CELL}*{lookup('F').format(item='Anchor Clamp - 48F Cable')}",
        "Anchor Clamp - 12F Cable": f"={POLE_CELL}*{lookup('F').format(item='Anchor Clamp - 12F Cable')}",
        "Universal Pole Bracket": f"={POLE_CELL}*{lookup('F').format(item='Universal Pole Bracket')}",
        "TespaBand": f"=MAX(1,ROUNDDOWN({POLE_CELL}/10,0)+1)",
        "Buckles": f"={POLE_CELL}*{lookup('F').format(item='Buckles')}",
        "Conduit (6m section)": (f"=IF({UG_CELL}>0,ROUNDDOWN({UG_CELL}/"
                                  f"{lookup('F').format(item='Conduit (6m section)')},0)+1,0)"),
        "Joint Box": f"={SEG_CELL}*{lookup('F').format(item='Joint Box')}",
        "Joint Box Installation": f"={SEG_CELL}*{lookup('F').format(item='Joint Box Installation')}",
        "Overhead Installation Labour": f"={OH_CELL}",
        "Underground Installation Labour": f"={UG_CELL}",
    }

    r = table_start + 1
    for item, qty_formula in items_qty.items():
        ce.cell(row=r, column=1, value=f'={lookup("A").format(item=item)}').border = BORDER
        ce.cell(row=r, column=2, value=f'={lookup("B").format(item=item)}').border = BORDER
        ce.cell(row=r, column=3, value=f'={lookup("C").format(item=item)}').border = BORDER
        pc = ce.cell(row=r, column=4, value=f'={lookup("D").format(item=item)}')
        pc.border = BORDER; pc.number_format = '#,##0'
        qc = ce.cell(row=r, column=5, value=f'={lookup("E").format(item=item)}')
        qc.border = BORDER; qc.font = Font(size=9, color="475569")
        qty = ce.cell(row=r, column=6, value=qty_formula)
        qty.border = BORDER; qty.number_format = '#,##0.0'
        sub = ce.cell(row=r, column=7, value=f"=D{r}*F{r}")
        sub.border = BORDER; sub.number_format = '#,##0'
        r += 1

    last_item_row = r - 1
    r += 1
    sub_row = r
    ce.cell(row=r, column=6, value="Subtotal").font = Font(bold=True)
    ce.cell(row=r, column=6).fill = TOTAL_FILL
    sc = ce.cell(row=r, column=7, value=f"=SUM(G{table_start+1}:G{last_item_row})")
    sc.font = Font(bold=True); sc.fill = TOTAL_FILL; sc.number_format = '#,##0'
    r += 1
    ce.cell(row=r, column=6, value="VAT 15%").font = Font(size=10, color="475569")
    vc = ce.cell(row=r, column=7, value=f"=G{sub_row}*0.15")
    vc.number_format = '#,##0'
    r += 1
    ce.cell(row=r, column=6, value="GRAND TOTAL (Rs)").font = Font(bold=True, color="FFFFFF")
    ce.cell(row=r, column=6).fill = GRAND_FILL
    gc = ce.cell(row=r, column=7, value=f"=G{sub_row}+G{sub_row+1}")
    gc.font = Font(bold=True, color="FFFFFF", size=12)
    gc.fill = GRAND_FILL
    gc.number_format = '#,##0'

    for i, w in enumerate([28, 11, 8, 14, 30, 12, 14], start=1):
        ce.column_dimensions[get_column_letter(i)].width = w
    ce.freeze_panes = f"A{table_start+1}"
    return ce


def _build_pole_list_sheet(wb, poles):
    ws = wb.create_sheet("Pole List")
    headers = ["pole_no", "distance_m", "percentage", "latitude", "longitude", "section_type"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER
    for r, pole in enumerate(poles, start=2):
        for i, h in enumerate(headers, start=1):
            ws.cell(row=r, column=i, value=pole.get(h)).border = BORDER
    for i, w in enumerate([10, 12, 11, 12, 12, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    return ws


def build_costing_workbook(oh_m, ug_m, seg_count, poles):
    """Build the InventoryA-referenced costing workbook for a computed route."""
    wb = Workbook()
    _build_inventory_sheet(wb)
    _build_cost_estimate_sheet(wb, oh_m, ug_m, seg_count)
    _build_pole_list_sheet(wb, poles)
    wb._sheets = [wb["InventoryA"], wb["Cost Estimate"], wb["Pole List"]]
    wb.active = 1
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ---------------------------------------------------------------------------
# Materials Schedule Download
# ---------------------------------------------------------------------------
@app.route("/api/download-xlsx", methods=["POST"])
def download_xlsx():
    data = request.get_json(force=True)

    route_geom = data.get("road_geometry", [])
    if len(route_geom) < 2:
        return jsonify({"error": "Invalid route geometry."}), 400

    ug_segments = data.get("ug_segments", [])

    # Generate poles and summary (still used for the Pole List sheet + route distances)
    poles, equipment_list, summary = generate_materials_schedule(route_geom, ug_segments)

    output = build_costing_workbook(
        oh_m=summary["oh_distance_m"],
        ug_m=summary["ug_distance_m"],
        seg_count=summary["ug_segment_count"],
        poles=poles,
    )

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"route_materials_{route_geom[0][0]:.4f}_{route_geom[-1][0]:.4f}.xlsx"
    )


@app.route('/api/<path:path>', methods=['OPTIONS'])
def handle_options(path):
    return '', 200


if __name__ == "__main__":
    print("=" * 50)
    print("Mauritius Route Planner API with Ollama AI")
    print("=" * 50)
    print(f"Stations loaded: {len(STATIONS)}")
    for s in STATIONS:
        print(f"  - {s['name']}: {s['lat']}, {s['lon']}")
    print("=" * 50)
    print(f"🤖 AI Model: {OLLAMA_MODEL}")
    try:
        import ollama
        print("✅ Ollama: Installed")
    except ImportError:
        print("❌ Ollama: Not installed (pip install ollama)")
    print("=" * 50)
    print("Starting server on http://localhost:5000")
    print("Test the API at: http://localhost:5000/api/stations")
    print("=" * 50)
    app.run(debug=True, port=5000)